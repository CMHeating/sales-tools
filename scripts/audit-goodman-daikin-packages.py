from pathlib import Path
import re, csv
from openpyxl import load_workbook

HTML = Path("hca-caas-quoting.html")
PRICEBOOK = Path("data/vendor-pricing/reliance-goodman-daikin-pricebook.xlsx")
OUT = Path("data/audits/goodman-daikin-package-audit.csv")

html = HTML.read_text()
wb = load_workbook(PRICEBOOK, data_only=True)
ws = wb["Customer Copy"]

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
idx = {h: i + 1 for i, h in enumerate(headers)}

price_rows = []
for r in range(2, ws.max_row + 1):
    brand = ws.cell(r, idx["Brand"]).value
    status = ws.cell(r, idx["Status"]).value
    product = ws.cell(r, idx["Product#"]).value
    price = ws.cell(r, idx["Price"]).value
    category = ws.cell(r, idx["Category"]).value
    desc = ws.cell(r, idx["Description"]).value
    line = ws.cell(r, idx["Product Line"]).value

    if not brand or brand not in ("Goodman", "Daikin"):
        continue
    if status and str(status).strip().lower() not in ("active", "a", "stock", "available"):
        continue

    price_rows.append({
        "brand": brand,
        "product": str(product).strip() if product else "",
        "price": price,
        "category": category,
        "description": desc,
        "line": line,
        "status": status,
    })

price_lookup = {x["product"]: x for x in price_rows if x["product"]}

pkg_re = re.compile(
    r'\{id:"(?P<id>[^"]+)",t:"(?P<type>[^"]+)",brand:"(?P<brand>Goodman|Daikin)",tier:"(?P<tier>[^"]+)",lbl:"(?P<label>[^"]+)".*?sz:\{(?P<sizes>.*?)\}\}',
    re.S
)

size_re = re.compile(r'"(?P<size>[^"]+)":\{mat:(?P<mat>[0-9.]+),msrp:(?P<msrp>[0-9.]+)\}')

rows = []
for m in pkg_re.finditer(html):
    pkg = m.groupdict()

    # Best-effort model match: current quote rows mostly store labels, not explicit model list.
    possible_model_hits = []
    label_blob = " ".join([pkg["id"], pkg["label"]]).upper()
    for product, pr in price_lookup.items():
        if product and product.upper() in label_blob:
            possible_model_hits.append(product)

    for sm in size_re.finditer(pkg["sizes"]):
        rows.append({
            "Package ID": pkg["id"],
            "Brand": pkg["brand"],
            "Type": pkg["type"],
            "Tier": pkg["tier"],
            "Label": pkg["label"],
            "Size": sm.group("size"),
            "Current Mat": float(sm.group("mat")),
            "Current MSRP": float(sm.group("msrp")),
            "Matched Product#": ", ".join(possible_model_hits),
            "Pricebook Match Count": len(possible_model_hits),
            "Audit Result": "NEEDS MODEL MAPPING" if len(possible_model_hits) == 0 else "HAS POSSIBLE MATCH",
        })

OUT.parent.mkdir(parents=True, exist_ok=True)
with OUT.open("w", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()) if rows else [
        "Package ID","Brand","Type","Tier","Label","Size","Current Mat","Current MSRP","Matched Product#","Pricebook Match Count","Audit Result"
    ])
    writer.writeheader()
    writer.writerows(rows)

print(f"Wrote {OUT}")
print(f"Pricebook active Goodman/Daikin rows: {len(price_rows)}")
print(f"Quote package size rows audited: {len(rows)}")
print("Next: open the CSV and identify packages marked NEEDS MODEL MAPPING.")
